"""
Модуль класса анализатора временного ряда и функции записи DataFrame 
в excel документ.
"""
import statsmodels.api as sm
import pandas as pd
from openpyxl import load_workbook, Workbook
from typing import Iterable

class Analyzer:
    """
    Класс, использующийся для вычисления дифференциала временного ряда,
    скользящего среднего, нахождения автокорреляции и точек экстремума.
    """
    def __init__(self, df: pd.DataFrame):
        """
        Конструктор для создания экземпляра класса Analyzer.
        Аргумент:
            df: временной ряд.
        """
        if not self._validate_init(df):
            raise ValueError(f"Analyzer build error\n"
                                f"\"df\": {df},\n")
        self._data_frame = df

    @staticmethod
    def _validate_init(df: pd.DataFrame) -> bool:
        """
        Метод валидации данных.

        Аргумент:
            df: временной ряд.

        Возвращает: 
            Булево значение результата валидации данных.
        """
        if not isinstance(df, pd.DataFrame):
            return False
        return True

    @property
    def data_frame(self) -> pd.DataFrame:
        """
        Метод получения DataFrame временного ряда.

        Возвращает:
            DataFrame временного ряда.
        """
        return self._data_frame
    
    @data_frame.setter
    def data_frame(self, df: pd.DataFrame) -> None:
        """
        Метод изменения значения DataFrame временного ряда.

        Аргумент:
            df: DataFrame временного ряда.

        Возвращает:
            Ничего не возвращает.
        """
        self.data_frame = df

    def _validate_rolling_meaning(method): 
        """
        Метод-декоратор для валидации данных, используемых методом,
        вычисляющим скользящее среднее.

        Аргумент:
            method: метод класса Analyzer, используемый для 
            преобразования DataFrame.

        Возвращает: 
            Функция, указанная в параметрах.
        """
        def wrapper(*args, **kwargs):
            temp_df = method(*args, **kwargs)
            if (len(args) != 2):
                raise TypeError(f"Incorrect arguments count\n"
                                f"\"arguments count\": {len(args)}.")
            if (not isinstance (args[1], int)):
                raise TypeError("Argument must be integer\n"
                                f"\"argument type\": {type(args[1])}.")
            if (args[1] < 1):
                raise ValueError("Argument must be > 0\n"
                                f"\"argument value\": {args[1]}.")
            return temp_df
        return wrapper

    @_validate_rolling_meaning
    def get_rolling_meaning(self, window_size: int) -> pd.DataFrame:
        """
            Метод вычисления скользящего среднего временного ряда с указанным
            размером окна.

            Аргумент:
                window_size: размер окна.

            Возвращает: 
                Скользящее среднее временного ряда.
        """
        temp_data = {}
        for i in self._data_frame.columns:
            temp_data[i] = self._data_frame[i].rolling(
                                                window = window_size).mean()
        new_df = pd.DataFrame(temp_data, index = self._data_frame.index)
        return new_df
    
    def get_differential(self) -> pd.DataFrame:
        """
        Метод вычисления дифференциала временного ряда.

        Возвращает:
            Дифференциал временного ряда.
        """
        new_df = self.data_frame.diff()
        return new_df
    
    def get_autocorrelation(self) -> pd.DataFrame:
        """
        Метод нахождения автокорелляции временного ряда.

        Возвращает: 
            Автокорреляция временного ряда.
        """
        for i in self.data_frame.columns:
            temp_list = []
            for j in range(len(self.data_frame.index)):
                temp_list.append(((self.data_frame-self.data_frame.mean()) *
                                (self.data_frame.shift(j)
                                - self.data_frame.mean()))
                                .mean() * (len(self.data_frame) - j)
                                / len(self.data_frame)
                                / self.data_frame.var(ddof = 0))

        temp_dict = {}
        temp_data = []
        for i in self.data_frame.columns:
            for j in range(len(self.data_frame.index)):
                temp_data.append(temp_list[j][i])
            temp_dict[i] = temp_data
            temp_data = []

        new_df = pd.DataFrame(temp_dict, index = self.data_frame.index)

        return new_df
    def get_extremums(self) -> pd.DataFrame:
        """
        Метод нахождения точек экстремума.
        
        Возвращает: 
            Точки экстремума.
        """
        diff = self.data_frame.diff()
        extremum_points_indexes = {}
        for i in self.get_keys():
            extremum_points_indexes[i] = []
            for j in range (1, len(diff[i])):
                if diff[i].iloc[j-1] < 0 and diff[i].iloc[j] > 0 or \
                    diff[i].iloc[j-1] > 0 and diff[i].iloc[j] < 0:
                    extremum_points_indexes[i].append(
                                            self.data_frame[i].index[j-1] 
                                            if abs(diff[i].iloc[j-1]) <=
                                                abs(diff[i].iloc[j]) 
                                            else self.data_frame[i].index[j])
                elif diff[i].iloc[j-1] == 0:
                    extremum_points_indexes[i].append(
                        self.data_frame[i].index[j-1])
                elif diff[i].iloc[j] == 0:
                    extremum_points_indexes[i].append(
                        self.data_frame[i].index[j])
        extremum_points_indexes_unique = {}
        for i in self.get_keys():
            extremum_points_indexes_unique[i] = []
            for j in extremum_points_indexes[i]:
                if j not in extremum_points_indexes_unique[i]:
                    extremum_points_indexes_unique[i].append(j)
        extremum_points = {}
        for i in self.get_keys():
            extremum_points[i] = []
            for j in extremum_points_indexes_unique[i]:
                extremum_points[i].append(self.data_frame[i][j])

        series_list = []
        for i in self.get_keys():
            series_list.append(pd.Series(extremum_points[i], 
                                         index = 
                                         extremum_points_indexes_unique[i]))
        new_df = {}
        for i in range(len(self.data_frame.columns)):
            new_df[self.data_frame.columns[i]] = series_list[i]
        new_df = pd.DataFrame(new_df)
        return new_df
    
    def get_keys(self) -> Iterable[str]:
        """
        Метод-генератор для поочередного вывода названий столбцов из
        DataFrame временного ряда.
        
        Возвращает: 
            Название столбца из DataFrame временного ряда.
        """
        for key in dict(self._data_frame).keys():
            yield key

    def get_cols(self) -> Iterable[pd.Series]:
        """
        Метод-генератор для поочередного вывода столбцов данных из
        DataFrame временного ряда.
        
        Возвращает: 
            Столбец данных из DataFrame временного ряда.
        """
        for key in dict(self._data_frame).keys():
            yield self.data_frame[key]

def write_in_excel(df: pd.DataFrame, title: str) -> None:
    """
    Функция, записывающая DataFrame в Excel файл.

    Аргумент:
        df: временной ряд.
        title: название листа в Excel-документе, в который произведётся запись.

    Возвращает:
        Ничего не возвращает.
    """
    try:
        file = load_workbook("data.xlsx")
        file.close()
        with pd.ExcelWriter("data.xlsx", mode="a", engine="openpyxl",
                            if_sheet_exists='new') as writer:
            df.to_excel(writer, sheet_name = title)
    except (FileNotFoundError, IndexError):
        file = Workbook()
        file.save('data.xlsx')
        file.close()
        with pd.ExcelWriter("data.xlsx", engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name = title)

if __name__ == '__main__':
    pass